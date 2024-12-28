from openpyxl import load_workbook
import pandas as pd
from standardize import initialize_AE, initialize_TD
from combine_and_split import combine_and_split_by_month
from write import write_to_master

if __name__ == "__main__":
    ae_df = initialize_AE("sheets/Summary.xls")
    td_credit_df = initialize_TD("sheets/accountactivity1.csv")
    td_debit_df = initialize_TD("sheets/accountactivity.csv", is_debit=True)
    split_df = combine_and_split_by_month(ae_df, td_credit_df, td_debit_df)
    write_to_master(split_df)


    # book = load_workbook('sheets/master.xlsx')
    # writer = pd.ExcelWriter('sheets/master.xlsx', engine='openpyxl')
    # # writer.book = book

    # for month, sub_df in split_df.items():
    #     print(f"Data for {month}:\n{sub_df}\n")
    #     sub_df.to_excel(writer, sheet_name=month, startrow=writer.sheets[month].max_row, index=False, header=False)


    # writer.save()

    # Things to do:
    # Need to resize columns again
    # Need to restructure code again

# Theory craft for removing duplicates
# Pull existing info from sheets into a df
# combine that df with the new one
# df = df.drop_duplicates(subset=['Description', 'Amount'], keep='first')
# Creates table that drops duplicate
# Write that to the sheet